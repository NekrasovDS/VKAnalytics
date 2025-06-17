import requests
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tkinter as tk
from tkinter import messagebox

count = 1

wb = Workbook()
ws = wb.active
ws.title = "VK Users"

headers = ['Имя', 'Фамилия', 'Id пользователя', 'Id группы', 'Направление']
ws.append(headers)

header_font = Font(bold=True, size=13, color="000000")
header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

def get_info_login():
    def on_submit():
        uid_input = entry_user_id.get().strip()
        token_input = entry_token.get().strip()

        if not uid_input or not token_input:
            messagebox.showerror("Ошибка", "Пожалуйста, введите и ID пользователя, и токен.")
            return

        try:
            uid = int(uid_input)
        except ValueError:
            messagebox.showerror("Ошибка", "ID пользователя должен быть числом.")
            return

        root.destroy()  # Закрыть окно
        get_info_params(uid, token_input)  # Вызов функции с аргументами

    root = tk.Tk()
    root.title("VK Информация")
    root.geometry("350x180")
    root.resizable(False, False)

    tk.Label(root, text="Введите ID пользователя:", anchor="w").pack(padx=10, pady=(10, 0), fill="x")
    entry_user_id = tk.Entry(root)
    entry_user_id.pack(padx=10, fill="x")

    tk.Label(root, text="Введите токен:", anchor="w").pack(padx=10, pady=(10, 0), fill="x")
    entry_token = tk.Entry(root)
    entry_token.pack(padx=10, fill="x")

    tk.Button(root, text="Далее", command=on_submit).pack(pady=15)

    root.mainloop()

for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = header_font
    ws.cell(row=1, column=col).fill = header_fill
    ws.column_dimensions[chr(64 + col)].width = 25

def get_info_params(uid, token_input):

    def save_parameters():
        global user_count, group_count
        try:
            user_count = int(entry_user_count.get())
            group_count = int(entry_group_count.get())

            messagebox.showinfo("Успех",
                                f"Параметры успешно сохранены!\nКоличество людей: {user_count}\nКоличество групп: {group_count}")
            window.destroy()
            get_info(uid, token_input)
        except ValueError:
            messagebox.showerror("Ошибка", "Пожалуйста, введите числовые значения.")

    # Создаем окно для ввода параметров
    window = tk.Tk()
    window.title("Параметры кластеризации")

    label_user_count = tk.Label(window, text="Количество пользователей:")
    label_user_count.pack(pady=5)
    entry_user_count = tk.Entry(window)
    entry_user_count.pack(pady=5)

    label_group_count = tk.Label(window, text="Количество групп:")
    label_group_count.pack(pady=5)
    entry_group_count = tk.Entry(window)
    entry_group_count.pack(pady=5)

    button_ok = tk.Button(window, text="Запустить", command=save_parameters)
    button_ok.pack(pady=20)

    window.mainloop()

def get_info(user_id, token):
    row_num = 2

    for i in range(user_count):
        print(f"\n🔁 ИТЕРАЦИЯ {i + 1}")
        params = {
            'user_id': user_id,
            'order': 'random',
            'count': count,
            'fields': 'first_name,last_name',
            'v': '5.199',
            'access_token': token
        }

        try:
            r = requests.get("https://api.vk.com/method/friends.get", params=params)
            data = r.json()
            if "error" in data:
                error_code = data['error'].get('error_code')
                error_msg = data['error'].get('error_msg', 'Неизвестная ошибка')

                # ⚠️ Если ошибка авторизации (ошибка 5 — "User authorization failed")
                if error_code == 5:
                    messagebox.showerror("⚠️ Ошибка авторизации", "Недействительные ID и/или токен.")
                    return  # остановить выполнение
                else:
                    print("Ошибка VK при получении друзей:", error_msg)
                    continue
            friends = data.get('response', {}).get('items', [])
        except Exception as e:
            print("Ошибка запроса друзей:", e)
            continue

        for friend in friends:
            print(f"\n👤 {friend['first_name']} {friend['last_name']} ({friend['id']})")
            ws.cell(row=row_num, column=1).value = friend['first_name']
            ws.cell(row=row_num, column=2).value = friend['last_name']
            ws.cell(row=row_num, column=3).value = friend['id']

            subs_params = {
                'user_id': friend['id'],
                'extended': 1,
                'count': group_count,
                'v': '5.199',
                'access_token': token
            }

            try:
                subs_r = requests.get("https://api.vk.com/method/users.getSubscriptions", params=subs_params)
                subs_data = subs_r.json()
                if "error" in subs_data:
                    print("Ошибка VK при получении подписок:", subs_data['error']['error_msg'])
                    continue
                groups = subs_data.get('response', {}).get('items', [])
            except Exception as e:
                print("Ошибка запроса подписок:", e)
                continue

            for group in groups:
                group_id = group.get('id')
                group_info_r = requests.get("https://api.vk.com/method/groups.getById", params={
                    'group_id': group_id,
                    'fields': 'activity',
                    'v': '5.199',
                    'access_token': token
                })

                group_data = group_info_r.json()
                if "error" in group_data:
                    print(f"⚠️ Ошибка VK для группы {group_id}: {group_data['error']['error_msg']}")
                    activity = "Ошибка VK"
                else:
                    groups = group_data.get('response', {}).get('groups', [])
                    if groups:
                        group_info = groups[0]
                        activity = group_info.get('activity', '—')
                    else:
                        activity = '—'

                time.sleep(1)
                print(f"📌 Группа {group_id}, тематика: {activity}")
                ws.cell(row=row_num, column=4).value = group_id
                ws.cell(row=row_num, column=5).value = activity
                row_num += 1
            row_num += 1
            time.sleep(2)

    wb.save("user_database_1.xlsx")
    print("\n✅ Данные сохранены в 'user_database_1.xlsx'")